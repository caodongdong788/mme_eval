from __future__ import annotations

from openpyxl import load_workbook

from medeval.reporter import build_report, write_transcripts_xlsx
from medeval.reporter.markdown_report import render_markdown
from tests.test_eight_dimension_scoring import result


def report():
    return build_report("v2", [result(guideline_score=2)], "stub")


def test_markdown_uses_45_point_eight_dimension_language() -> None:
    text = render_markdown(report())
    assert "八维与三端评分" in text
    assert "/45" in text
    assert "医学安全性" in text
    assert "指南得分率（缺分已扣入对应维度）" in text
    assert "四模块" not in text


def test_excel_contains_dimensions_ends_and_guidelines(tmp_path) -> None:
    path = write_transcripts_xlsx(report(), tmp_path / "v2.xlsx")
    sheet = load_workbook(path)["对话流水"]
    headers = [cell.value for cell in sheet[1]]
    for name in ("医学安全性", "专业准确性与边界", "医生端", "护士端", "患者端", "指南评分"):
        assert name in headers
    values = [cell.value for cell in sheet[2]]
    assert any(value == "44/45" for value in values)
    assert any(isinstance(value, str) and "risk: 2/3" in value for value in values)
    reason = sheet.cell(row=2, column=headers.index("扣分原因") + 1).value
    assert "professional_accuracy 指南 risk -1分：stub" in reason


def test_excel_contains_low_dimension_reason_without_formula_deduction(tmp_path) -> None:
    item = result(guideline_score=3)
    inquiry = next(
        verdict
        for verdict in item.verdicts
        if verdict.name == "dimension.clinical_inquiry"
    )
    inquiry.score = 2
    inquiry.reason = "未追问治疗阶段，无法针对当前情况给出建议。"

    path = write_transcripts_xlsx(
        build_report("v2", [item], "stub"),
        tmp_path / "low-dimension.xlsx",
    )
    sheet = load_workbook(path)["对话流水"]
    headers = [cell.value for cell in sheet[1]]
    reason = sheet.cell(row=2, column=headers.index("扣分原因") + 1).value

    assert item.score_deductions == []
    assert reason == (
        "临床追问充分性 2/5："
        "未追问治疗阶段，无法针对当前情况给出建议。"
    )
