from __future__ import annotations

import zipfile
from io import BytesIO

from openpyxl import load_workbook
from PIL import Image as PILImage

from server.db import session_scope
from server.models_db import OnlineEval, OnlineEvalCase
from server.services import online_eval_export as export_svc
from server.services.online_eval_export import case_dialogue_turns, split_filter_values


def _png_bytes(width: int = 100, height: int = 200, color=(200, 40, 40)) -> bytes:
    buf = BytesIO()
    PILImage.new("RGB", (width, height), color).save(buf, format="PNG")
    return buf.getvalue()


def _img_case(external_id: str, content: str, role: str = "doctor") -> OnlineEvalCase:
    return OnlineEvalCase(
        external_id=external_id,
        case_name=f"{external_id}-图片",
        user_text="",
        assistant_text="",
        raw_messages=[
            {"role": "user", "content": content},
            {"role": "assistant", "content": "报告解读回复"},
        ],
        gate_status="pass",
        total_score=42.0,
        grade="excellent",
        task_type="report_interpretation",
        review_role=role,
    )


def _media_names(path) -> list[str]:
    return [n for n in zipfile.ZipFile(path).namelist() if n.startswith("xl/media/")]


def _seed_online_eval(*, classified: bool = True) -> int:
    with session_scope() as session:
        row = OnlineEval(
            name="导出批次",
            status="success",
            case_count=2,
            avg_score=21.0,
        )
        row.cases.append(
            OnlineEvalCase(
                external_id="case-pass",
                case_name="骨密度复诊咨询",
                user_text="第一问\n第二问",
                assistant_text="第一答\n第二答",
                raw_messages=[
                    {"role": "user", "content": "第一问"},
                    {"role": "assistant", "content": "第一答"},
                    {"role": "user", "content": "第二问"},
                    {"role": "assistant", "content": "第二答"},
                ],
                gate_status="pass",
                total_score=42.0,
                grade="excellent",
                task_type="report_interpretation",
                review_role="doctor" if classified else "",
            )
        )
        row.cases.append(
            OnlineEvalCase(
                external_id="case-fail",
                case_name="停药建议",
                user_text="能不能停药",
                assistant_text="可以自行停药",
                raw_messages=[
                    {"role": "user", "content": "能不能停药"},
                    {"role": "assistant", "content": "可以自行停药"},
                ],
                gate_status="fail",
                total_score=0.0,
                grade="unqualified",
                task_type="adherence_side_effect",
                review_role="nurse" if classified else "",
            )
        )
        session.add(row)
        session.flush()
        return row.id


def test_online_eval_export_filters_and_writes_multiturn_xlsx(client, monkeypatch):
    eval_id = _seed_online_eval()
    captured: dict = {}

    def fake_publish(xlsx_path, *, parent_folder_token, title):
        captured["folder"] = parent_folder_token
        captured["title"] = title
        wb = load_workbook(xlsx_path)
        captured["sheetnames"] = wb.sheetnames
        ws = wb.active
        captured["rows"] = [tuple(row) for row in ws.iter_rows(values_only=True)]
        return "https://feishu.example/sheets/online-export"

    monkeypatch.setattr(
        "server.services.online_eval_export.publish_xlsx_to_lark", fake_publish
    )

    resp = client.post(
        f"/api/online-evals/{eval_id}/export-cases",
        params={
            "gate_status": "pass",
            "score_bucket": "gte40_5",
            "grade": "excellent",
            "parent_folder_token": "",
        },
    )

    assert resp.status_code == 200, resp.text
    assert resp.json()["url"] == "https://feishu.example/sheets/online-export"
    assert resp.json()["count"] == 1
    assert captured["folder"] == ""
    assert captured["title"] == "导出批次_评测清单"
    # 过滤后只剩医生角色的 case，故只生成「医生」一个 sheet。
    assert captured["sheetnames"] == ["医生"]
    assert captured["rows"][0] == (
        "第一轮用户输入",
        "第一轮Cx输出",
        "第二轮用户输入",
        "第二轮Cx输出",
    )
    assert captured["rows"][1] == ("第一问", "第一答", "第二问", "第二答")


def test_online_eval_export_classifies_missing_roles_and_splits_sheets(client, monkeypatch):
    eval_id = _seed_online_eval(classified=False)
    captured: dict = {}

    # mock 掉 LLM：judge 占位 + 分类直接按 external_id 派发角色。
    monkeypatch.setattr(export_svc, "_resolve_online_judge", lambda *a, **k: object())

    async def fake_classify(cases, judge):
        mapping = {"case-pass": "doctor", "case-fail": "nurse"}
        return {case.id: mapping.get(case.external_id, "patient") for case in cases}

    monkeypatch.setattr(export_svc, "classify_missing_roles", fake_classify)

    def fake_publish(xlsx_path, *, parent_folder_token, title):
        wb = load_workbook(xlsx_path)
        captured["sheetnames"] = wb.sheetnames
        captured["doctor_rows"] = [tuple(r) for r in wb["医生"].iter_rows(values_only=True)]
        captured["nurse_rows"] = [tuple(r) for r in wb["护士"].iter_rows(values_only=True)]
        return "https://feishu.example/sheets/split"

    monkeypatch.setattr(
        "server.services.online_eval_export.publish_xlsx_to_lark", fake_publish
    )

    resp = client.post(
        f"/api/online-evals/{eval_id}/export-cases",
        params={"parent_folder_token": ""},
    )

    assert resp.status_code == 200, resp.text
    assert resp.json()["count"] == 2
    # 两个角色各一个 sheet，医生在前、护士在后，无患者 sheet。
    assert captured["sheetnames"] == ["医生", "护士"]
    assert captured["doctor_rows"][1][0] == "第一问"
    assert captured["nurse_rows"][1][0] == "能不能停药"

    # 分类结果已落库，重复导出不会再触发分类。
    with session_scope() as session:
        roles = {
            case.external_id: case.review_role
            for case in session.get(OnlineEval, eval_id).cases
        }
    assert roles == {"case-pass": "doctor", "case-fail": "nurse"}


def test_online_eval_export_empty_filter_returns_400(client, monkeypatch):
    eval_id = _seed_online_eval()

    monkeypatch.setattr(
        "server.services.online_eval_export.publish_xlsx_to_lark",
        lambda *args, **kwargs: "https://feishu.example/sheets/unused",
    )

    resp = client.post(
        f"/api/online-evals/{eval_id}/export-cases",
        params={"gate_status": "pass", "grade": "unqualified"},
    )

    assert resp.status_code == 400
    assert "没有可导出" in resp.json()["detail"]


def test_case_dialogue_turns_falls_back_to_flat_text(session):
    case = OnlineEvalCase(
        user_text="单轮用户",
        assistant_text="单轮回答",
        raw_messages=[],
        gate_status="pass",
        total_score=39.0,
        grade="good",
    )

    assert case_dialogue_turns(case) == [("单轮用户", "单轮回答")]


def test_split_filter_values_trims_blanks():
    assert split_filter_values(" pass, fail ,,") == ["pass", "fail"]


def test_extract_cell_images_extracts_refs_and_clean_text():
    text = "[图片：image_token=TKA，尺寸=100x200]\n请看报告\n[图片：image_token=TKB]"
    clean, refs = export_svc._extract_cell_images(text)
    assert [ref.token for ref in refs] == ["TKA", "TKB"]
    assert refs[0].placeholder == "[图片：image_token=TKA，尺寸=100x200]"
    assert clean == "请看报告"
    # 无图文本原样返回。
    assert export_svc._extract_cell_images("普通文本") == ("普通文本", [])


def test_write_cases_xlsx_embeds_real_image(tmp_path):
    case = _img_case("case-img", "[图片：image_token=TKA，尺寸=100x200]")
    fetcher = lambda token: _png_bytes(1200, 1600)  # noqa: E731
    path = tmp_path / "embed.xlsx"

    export_svc._write_cases_xlsx([case], path, fetcher)

    media = _media_names(path)
    assert len(media) == 1  # 嵌入了一张真实图片
    embedded = PILImage.open(BytesIO(zipfile.ZipFile(path).read(media[0])))
    assert embedded.size == (1200, 1600)  # 单图保留原始分辨率，避免导入飞书后发糊
    wb = load_workbook(path)
    ws = wb["医生"]
    assert ws["A2"].value in ("", None)  # 图片占位文本已被移除
    assert ws.column_dimensions["A"].width and ws.column_dimensions["A"].width > 45
    assert ws.row_dimensions[2].height and ws.row_dimensions[2].height > 300  # 行高已撑起


def test_write_cases_xlsx_stacks_multiple_images(tmp_path):
    content = "[图片：image_token=TKA，尺寸=100x200][图片：image_token=TKB，尺寸=100x200]"
    case = _img_case("case-multi", content)
    fetcher = lambda token: _png_bytes(1200, 1600)  # noqa: E731
    path = tmp_path / "stack.xlsx"

    export_svc._write_cases_xlsx([case], path, fetcher)

    media = _media_names(path)
    assert len(media) == 1  # 两张图竖直拼接成一张
    img = PILImage.open(BytesIO(zipfile.ZipFile(path).read(media[0])))
    # 拼接图按展示尺寸的多倍分辨率导出，明显高于实际展示宽度。
    assert img.width > 700
    assert img.height > 2500


def test_write_cases_xlsx_splits_mixed_text_and_images_into_duplicate_user_columns(tmp_path):
    content = "请看今天的报告\n[图片：image_token=TKA，尺寸=100x200][图片：image_token=TKB，尺寸=100x200]"
    case = _img_case("case-mixed", content)
    fetcher = lambda token: _png_bytes(1200, 1600)  # noqa: E731
    path = tmp_path / "mixed.xlsx"

    export_svc._write_cases_xlsx([case], path, fetcher)

    wb = load_workbook(path)
    ws = wb["医生"]
    assert [cell.value for cell in ws[1]] == ["第一轮用户输入", "第一轮用户输入", "第一轮Cx输出"]
    assert ws["A2"].value == "请看今天的报告"
    assert ws["B2"].value in ("", None)
    assert ws["C2"].value == "报告解读回复"
    media = _media_names(path)
    assert len(media) == 1
    img = PILImage.open(BytesIO(zipfile.ZipFile(path).read(media[0])))
    assert img.height > 2500


def test_write_cases_xlsx_splits_mixed_text_and_image_placeholders_without_fetcher(tmp_path):
    content = "请看今天的报告\n[图片：image_token=TKA，尺寸=100x200]"
    case = _img_case("case-mixed-fallback", content)
    path = tmp_path / "mixed-fallback.xlsx"

    export_svc._write_cases_xlsx([case], path, None)

    ws = load_workbook(path)["医生"]
    assert [cell.value for cell in ws[1]] == ["第一轮用户输入", "第一轮用户输入", "第一轮Cx输出"]
    assert ws["A2"].value == "请看今天的报告"
    assert ws["B2"].value == "[图片：image_token=TKA，尺寸=100x200]"
    assert ws["C2"].value == "报告解读回复"


def test_write_cases_xlsx_adapts_text_rows_and_wraps_cells(tmp_path):
    long_text = "这是一段很长的线上用户问题，导出后应该自动换行并撑开行高。" * 8
    case = _img_case("case-text", long_text)
    path = tmp_path / "adaptive.xlsx"

    export_svc._write_cases_xlsx([case], path, None)

    ws = load_workbook(path)["医生"]
    assert ws["A2"].alignment.wrap_text is True
    assert ws["A2"].alignment.vertical == "top"
    assert ws.row_dimensions[2].height and ws.row_dimensions[2].height > 30
    assert ws.column_dimensions["A"].width and ws.column_dimensions["A"].width >= 32


def test_write_cases_xlsx_keeps_text_when_no_fetcher(tmp_path):
    content = "[图片：image_token=TKA，尺寸=100x200]"
    case = _img_case("case-fallback", content)
    path = tmp_path / "fallback.xlsx"

    export_svc._write_cases_xlsx([case], path, None)

    assert _media_names(path) == []  # 未嵌图
    ws = load_workbook(path)["医生"]
    assert ws["A2"].value == content  # 保留原占位文本兜底


def test_write_cases_xlsx_falls_back_when_download_fails(tmp_path):
    case = _img_case("case-dlfail", "[图片：image_token=TKA，尺寸=100x200]")
    path = tmp_path / "dlfail.xlsx"

    export_svc._write_cases_xlsx([case], path, lambda token: None)  # 下载失败

    assert _media_names(path) == []
    assert load_workbook(path)["医生"]["A2"].value == "[图片：image_token=TKA，尺寸=100x200]"


def test_export_end_to_end_embeds_images_for_logged_in_user(client, monkeypatch):
    from server.auth import get_current_user_optional
    from server.feishu_media import FeishuMedia
    from server.models_db import FeishuUser

    with session_scope() as session:
        row = OnlineEval(name="图片批次", status="success", case_count=1)
        row.cases.append(_img_case("case-e2e", "[图片：image_token=TKA，尺寸=100x200]"))
        session.add(row)
        session.flush()
        eval_id = row.id

    # 避免真实刷新 token / 真实下载 / 真实上传飞书。
    monkeypatch.setattr(export_svc, "ensure_fresh_token", lambda *a, **k: None)
    monkeypatch.setattr(
        export_svc,
        "fetch_media",
        lambda access_token, token: FeishuMedia(content=_png_bytes(), content_type="image/png"),
    )
    captured: dict = {}

    def fake_import(access_token, xlsx_path, *, folder_token, title):
        captured["media"] = _media_names(xlsx_path)
        return "https://feishu.example/sheets/e2e"

    monkeypatch.setattr(export_svc, "import_xlsx_as_sheet", fake_import)

    client.app.dependency_overrides[get_current_user_optional] = lambda: FeishuUser(
        open_id="ou_test", name="测试用户", access_token="u-token"
    )
    try:
        resp = client.post(
            f"/api/online-evals/{eval_id}/export-cases",
            params={"parent_folder_token": ""},
        )
    finally:
        client.app.dependency_overrides.clear()

    assert resp.status_code == 200, resp.text
    assert resp.json()["url"] == "https://feishu.example/sheets/e2e"
    assert len(captured["media"]) == 1  # 登录用户导出时图片被真实嵌入
