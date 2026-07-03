from __future__ import annotations

from openpyxl import load_workbook
from sqlalchemy import select

from server.db import session_scope
from server.models_db import (
    FeishuUser,
    OnlineAnnotationPoolCase,
    OnlineAnnotationPoolPath,
    OnlineEval,
    OnlineEvalCase,
)
from server.services import online_annotation_pool as pool_svc


def _seed_online_case() -> int:
    with session_scope() as session:
        row = OnlineEval(name="标注来源", status="success", case_count=1)
        row.cases.append(
            OnlineEvalCase(
                external_id="case-good",
                case_name="满意样本",
                user_text="第一问",
                assistant_text="第一答",
                raw_messages=[
                    {"role": "user", "content": "第一问"},
                    {"role": "assistant", "content": "第一答"},
                ],
                user_profile="年龄：36\n治疗阶段：内分泌治疗中",
                task_type="general_support",
                review_role="patient",
                gate_status="pass",
                total_score=39.0,
                grade="good",
            )
        )
        session.add(row)
        session.flush()
        return row.cases[0].id


def test_online_annotation_pool_adds_case_once(client):
    case_id = _seed_online_case()

    create_resp = client.post(
        "/api/online-annotation-pool/paths",
        json={"path": "骨健康/满意样本", "description": "高质量回复"},
    )
    assert create_resp.status_code == 201, create_resp.text
    path_id = create_resp.json()["id"]

    first = client.post(
        f"/api/online-annotation-pool/paths/{path_id}/cases",
        json={"online_eval_case_id": case_id},
    )
    second = client.post(
        f"/api/online-annotation-pool/paths/{path_id}/cases",
        json={"online_eval_case_id": case_id},
    )
    assert first.status_code == 201, first.text
    assert second.status_code == 201, second.text
    assert first.json()["id"] == second.json()["id"]
    assert first.json()["user_profile"] == "年龄：36\n治疗阶段：内分泌治疗中"

    paths = client.get("/api/online-annotation-pool/paths").json()
    assert paths[0]["path"] == "骨健康/满意样本"
    assert paths[0]["case_count"] == 1


def test_online_annotation_pool_normalises_case_name_and_grade(client):
    with session_scope() as session:
        path = OnlineAnnotationPoolPath(path="旧快照修正")
        session.add(path)
        session.flush()
        session.add(
            OnlineAnnotationPoolCase(
                path_id=path.id,
                source_eval_id=1,
                source_case_id=1,
                external_id="old-case",
                case_name="旧快照名称",
                user_text="第一句用户问话\n第二句用户问话",
                assistant_text="答复",
                raw_messages=[
                    {"role": "user", "content": "真正第一句患者问话"},
                    {"role": "assistant", "content": "答复"},
                    {"role": "user", "content": "后续追问"},
                ],
                gate_status="pass",
                total_score=0.0,
                grade="high_quality",
            )
        )
        session.flush()
        path_id = path.id

    cases = client.get(f"/api/online-annotation-pool/paths/{path_id}/cases").json()

    assert cases[0]["case_name"] == "真正第一句患者问话"
    assert cases[0]["grade"] == "unqualified"


def test_online_annotation_pool_updates_path(client):
    case_id = _seed_online_case()
    path_id = client.post(
        "/api/online-annotation-pool/paths",
        json={"path": "  0703标注集  ", "description": "旧描述"},
    ).json()["id"]
    client.post(
        f"/api/online-annotation-pool/paths/{path_id}/cases",
        json={"online_eval_case_id": case_id},
    )

    resp = client.patch(
        f"/api/online-annotation-pool/paths/{path_id}",
        json={"path": "0704标注集", "description": "新描述"},
    )

    assert resp.status_code == 200, resp.text
    data = resp.json()
    assert data["path"] == "0704标注集"
    assert data["description"] == "新描述"
    assert data["case_count"] == 1

    paths = client.get("/api/online-annotation-pool/paths").json()
    assert paths == [data]


def test_online_annotation_pool_update_rejects_duplicate_path(client):
    first_id = client.post(
        "/api/online-annotation-pool/paths",
        json={"path": "标注集A"},
    ).json()["id"]
    client.post(
        "/api/online-annotation-pool/paths",
        json={"path": "标注集B"},
    )

    resp = client.patch(
        f"/api/online-annotation-pool/paths/{first_id}",
        json={"path": "标注集B", "description": ""},
    )

    assert resp.status_code == 409
    assert "已存在" in resp.json()["detail"]


def test_online_annotation_pool_imports_all_feishu_sheet_tabs(initialized_db, monkeypatch):
    def fake_fetch(access_token, source_url):
        assert access_token == "user-token"
        assert "VYOPsjT8zhIcN0tEvQXc6jernSg" in source_url
        headers = ["第一轮用户输入", "第一轮Cx输出", "第二轮用户输入", "第二轮Cx输出"]
        return [
            {
                "sheet_id": "0XlJjc",
                "sheet_name": "医生",
                "row_indices": [1, 2],
                "cells": [headers, ["医生首问", "医生首答", "", ""]],
            },
            {
                "sheet_id": "1fkcwr",
                "sheet_name": "护士",
                "row_indices": [1, 2],
                "cells": [headers, ["护士首问", "护士首答", "护士追问", "护士追答"]],
            },
        ]

    monkeypatch.setattr(pool_svc, "ensure_fresh_token", lambda session, user, settings: user)
    monkeypatch.setattr(pool_svc.feishu_sheet, "fetch_sheet_cells", fake_fetch)

    user = FeishuUser(open_id="ou_test", name="冬东", access_token="user-token")
    with session_scope() as session:
        pool_path = pool_svc.import_path_from_feishu_url(
            session,
            path="0703标注集",
            description="飞书导入",
            source_url="https://p130box8iy5.feishu.cn/sheets/VYOPsjT8zhIcN0tEvQXc6jernSg?sheet=0XlJjc",
            current_user=user,
        )
        cases = list(
            session.scalars(
                select(OnlineAnnotationPoolCase)
                .where(OnlineAnnotationPoolCase.path_id == pool_path.id)
                .order_by(OnlineAnnotationPoolCase.id.asc())
            )
        )

    assert pool_path.case_count == 2
    assert [case.review_role for case in cases] == ["doctor", "nurse"]
    assert cases[0].user_text == "医生首问"
    assert cases[1].raw_messages == [
        {"role": "user", "content": "护士首问"},
        {"role": "assistant", "content": "护士首答"},
        {"role": "user", "content": "护士追问"},
        {"role": "assistant", "content": "护士追答"},
    ]
    assert cases[0].source_case_id < 0
    assert cases[1].source_case_id < 0


def test_online_annotation_pool_import_requires_feishu_login(client):
    resp = client.post(
        "/api/online-annotation-pool/paths/import-feishu",
        json={
            "path": "未登录导入",
            "description": "",
            "source_url": "https://p130box8iy5.feishu.cn/sheets/VYOPsjT8zhIcN0tEvQXc6jernSg",
        },
    )

    assert resp.status_code == 401


def test_online_annotation_pool_deletes_path_and_cases(client):
    case_id = _seed_online_case()
    path_id = client.post(
        "/api/online-annotation-pool/paths",
        json={"path": "待删除标注集"},
    ).json()["id"]
    client.post(
        f"/api/online-annotation-pool/paths/{path_id}/cases",
        json={"online_eval_case_id": case_id},
    )

    resp = client.delete(f"/api/online-annotation-pool/paths/{path_id}")

    assert resp.status_code == 204, resp.text
    assert client.get("/api/online-annotation-pool/paths").json() == []
    assert client.get(f"/api/online-annotation-pool/paths/{path_id}/cases").status_code == 404
    assert client.delete(f"/api/online-annotation-pool/paths/{path_id}").status_code == 404


def test_online_annotation_pool_deletes_single_case(client):
    case_id = _seed_online_case()
    path_id = client.post(
        "/api/online-annotation-pool/paths",
        json={"path": "单条删除标注集"},
    ).json()["id"]
    added = client.post(
        f"/api/online-annotation-pool/paths/{path_id}/cases",
        json={"online_eval_case_id": case_id},
    ).json()

    resp = client.delete(f"/api/online-annotation-pool/paths/{path_id}/cases/{added['id']}")

    assert resp.status_code == 204, resp.text
    assert client.get(f"/api/online-annotation-pool/paths/{path_id}/cases").json() == []
    assert client.get("/api/online-annotation-pool/paths").json()[0]["case_count"] == 0
    assert client.delete(f"/api/online-annotation-pool/paths/{path_id}/cases/{added['id']}").status_code == 404


def test_online_annotation_pool_exports_path_cases(client, monkeypatch):
    case_id = _seed_online_case()
    path_id = client.post(
        "/api/online-annotation-pool/paths",
        json={"path": "导出路径"},
    ).json()["id"]
    client.post(
        f"/api/online-annotation-pool/paths/{path_id}/cases",
        json={"online_eval_case_id": case_id},
    )
    captured = {}

    def fake_publish(xlsx_path, *, parent_folder_token, title):
        wb = load_workbook(xlsx_path)
        ws = wb["患者"]
        captured["title"] = title
        captured["rows"] = [tuple(row) for row in ws.iter_rows(values_only=True)]
        return "https://feishu.example/sheets/pool"

    monkeypatch.setattr(
        "server.services.online_annotation_pool.publish_xlsx_to_lark",
        fake_publish,
    )

    resp = client.post(
        f"/api/online-annotation-pool/paths/{path_id}/export-cases",
        params={"parent_folder_token": ""},
    )

    assert resp.status_code == 200, resp.text
    assert resp.json()["count"] == 1
    assert resp.json()["url"] == "https://feishu.example/sheets/pool"
    assert captured["title"] == "导出路径_标注池清单"
    assert captured["rows"][0] == ("第一轮用户输入", "第一轮Cx输出")
    assert captured["rows"][1] == ("第一问", "第一答")
