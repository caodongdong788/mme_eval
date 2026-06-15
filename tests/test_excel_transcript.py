"""Test transcripts.xlsx writer (change add-transcript-excel-output)."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from medeval.models import (
    CaseResult,
    ChatMessage,
    ConversationTrace,
    JudgeVerdict,
    Level,
    RunReport,
    ScoreProfile,
    TestCase,
    Turn,
)
from medeval.reporter import write_transcripts_xlsx


def _case(sid: str, depth: int = 2) -> TestCase:
    return TestCase(
        sample_id=sid,
        scenario="测试场景",
        level=Level.L2,
        turns=[Turn(role="user", content=f"q{i+1}") for i in range(depth)],
    )


def _result(case: TestCase, *, passed: bool, stability: str = "stable_pass") -> CaseResult:
    msgs = [ChatMessage(role="system", content="你是医疗 bot")]
    for i in range(2):
        msgs.append(ChatMessage(role="user", content=f"用户问题 {i+1}"))
        msgs.append(ChatMessage(role="assistant", content=f"Bot 回答 {i+1}\n第二段长文本"))
    return CaseResult(
        case=case,
        trace=ConversationTrace(messages=msgs),
        verdicts=[],
        hard_gate_passed=passed,
        gate_passed=passed,
        release_passed=passed,
        failure_tags=[] if passed else ["missed_red_flag"],
        stability=stability,  # type: ignore[arg-type]
        n_runs=1 if stability != "flaky" else 3,
        per_run_gate_passed=[passed] if stability != "flaky" else [True, True, False],
        started_at=datetime.utcnow(),
        finished_at=datetime.utcnow(),
    )


def _make_report(n_cases: int) -> RunReport:
    results = []
    for i in range(n_cases):
        c = _case(f"case_{i:02d}", depth=2)
        results.append(_result(c, passed=(i % 2 == 0)))
    return RunReport(run_name="t", results=results, total=n_cases, n_runs=1)


def test_write_xlsx_basic_structure(tmp_path: Path):
    report = _make_report(5)
    path = tmp_path / "transcripts.xlsx"
    write_transcripts_xlsx(report, path)
    assert path.exists()

    wb = load_workbook(path)
    assert wb.sheetnames == ["概览", "对话流水"]


def test_overview_sheet_columns_and_rows(tmp_path: Path):
    report = _make_report(3)
    path = tmp_path / "transcripts.xlsx"
    write_transcripts_xlsx(report, path)
    wb = load_workbook(path)
    ov = wb["概览"]

    headers = [c.value for c in ov[1]]
    assert headers == [
        "sample_id",
        "level",
        "depth",
        "scenario",
        "passed",
        "stability",
        "failure_tags",
        "评分档",
    ]
    # 3 case + 1 header
    assert ov.max_row == 4
    # 第一行 data: case_00 / L2 / depth=2 / 测试场景 / passed=True / stable_pass / ""
    row2 = [c.value for c in ov[2]]
    assert row2[0] == "case_00"
    assert row2[1] == "L2"
    assert row2[2] == 2  # depth = user turn count
    assert row2[3] == "测试场景"
    assert row2[4] is True
    assert row2[5] == "stable_pass"


def test_transcripts_sheet_one_row_per_case(tmp_path: Path):
    report = _make_report(2)
    path = tmp_path / "transcripts.xlsx"
    write_transcripts_xlsx(report, path)
    wb = load_workbook(path)
    tr = wb["对话流水"]

    # 宽表：前缀列 + 每轮 内容/耗时 两列（本用例最大 2 轮）
    headers = [c.value for c in tr[1]]
    assert headers == [
        "测试内容",
        "安全",
        "合规",
        "功能",
        "体验",
        "总分",
        "评级",
        "扣分原因",
        "得分点净分",
        "指南匹配率",
        "得分点明细",
        "轮数",
        "总耗时(ms)",
        "第1轮（用户+Bot）",
        "第1轮耗时(ms)",
        "第2轮（用户+Bot）",
        "第2轮耗时(ms)",
    ]

    # 每条 case 占 1 行：2 case + 1 header = 3 行
    assert tr.max_row == 3

    # case_00 行：测试内容=scenario(无 sub_scenario 时回退) / 轮数=2
    row2 = [c.value for c in tr[2]]
    assert row2[0] == "测试场景\ndefault（默认）"  # 无 case_file 注入时仅描述 + profile
    assert row2[11] == 2           # 轮数列
    # 第 1 轮内容 cell（第 14 列）同时含用户与 bot 两块
    c1 = row2[13]
    assert "👤 用户：用户问题 1" in c1
    assert "🤖 Bot：Bot 回答 1" in c1


def test_test_content_includes_profile_line(tmp_path: Path):
    """测试内容列含来源文件名与 profile（英文 + 中文）。"""
    case = TestCase(
        sample_id="adv",
        scenario="对抗",
        sub_scenario="术后腰痛与骨转移担忧",
        level=Level.L4,
        score_profile=ScoreProfile.adversarial,
        turns=[Turn(role="user", content="q")],
        case_file="adversarial.yaml",
    )
    r = CaseResult(
        case=case,
        trace=ConversationTrace(
            messages=[
                ChatMessage(role="user", content="q"),
                ChatMessage(role="assistant", content="a"),
            ]
        ),
        verdicts=[],
        hard_gate_passed=True,
        release_passed=True,
    )
    scoring = {
        "profiles": {
            "adversarial": {
                "module_max": {
                    "safety": 0.45,
                    "compliance": 0.20,
                    "function": 0.25,
                    "experience": 0.10,
                },
            },
        },
        "profile_match": [
            {
                "when": {"tags_any": ["adversarial"], "level_any": ["L4"]},
                "profile": "adversarial",
            },
        ],
    }
    report = RunReport(
        run_name="t",
        results=[r],
        total=1,
        n_runs=1,
        config_snapshot={"scoring": scoring},
    )
    path = tmp_path / "transcripts.xlsx"
    write_transcripts_xlsx(report, path)
    tr = load_workbook(path)["对话流水"]
    assert tr.cell(2, 1).value == (
        "术后腰痛与骨转移担忧\nadversarial.yaml\nadversarial（对抗 / 干扰）"
    )


def test_test_label_prefers_sub_scenario(tmp_path: Path):
    case = TestCase(
        sample_id="sid",
        scenario="筛查场景",
        sub_scenario="BI-RADS 3 类随访",
        level=Level.L2,
        turns=[Turn(role="user", content="q")],
    )
    r = _result(case, passed=True)
    report = RunReport(run_name="t", results=[r], total=1, n_runs=1)
    path = tmp_path / "transcripts.xlsx"
    write_transcripts_xlsx(report, path)
    tr = load_workbook(path)["对话流水"]
    assert tr.cell(2, 1).value == "BI-RADS 3 类随访\ndefault（默认）"


def test_deduction_reason_column(tmp_path: Path):
    """score_deductions 逐条落到「扣分原因」列（第 8 列）；空则为 —。"""
    case = _case("failcase", depth=1)
    r = CaseResult(
        case=case,
        trace=ConversationTrace(
            messages=[
                ChatMessage(role="user", content="问"),
                ChatMessage(role="assistant", content="答"),
            ]
        ),
        verdicts=[],
        hard_gate_passed=True,
        release_passed=False,
        score_deductions=[
            "安全 -0.30：未给出明确就医建议",
            "功能 -0.10：命中 must_not_have「马上手术」",
        ],
    )
    report = RunReport(run_name="t", results=[r], total=1, n_runs=1)
    path = tmp_path / "transcripts.xlsx"
    write_transcripts_xlsx(report, path)
    tr = load_workbook(path)["对话流水"]
    reason = tr.cell(2, 8).value  # 扣分原因列
    assert "安全 -0.30" in reason
    assert "命中 must_not_have「马上手术」" in reason


def test_passed_case_deduction_dash(tmp_path: Path):
    report = _make_report(1)  # case_00 无 score_deductions
    path = tmp_path / "transcripts.xlsx"
    write_transcripts_xlsx(report, path)
    tr = load_workbook(path)["对话流水"]
    assert tr.cell(2, 8).value == "—"


def _highlight_case() -> CaseResult:
    return CaseResult(
        case=_case("hl", depth=1),
        trace=ConversationTrace(
            messages=[
                ChatMessage(role="user", content="要不要马上手术？"),
                ChatMessage(role="assistant", content="建议先随访，不需要马上手术。"),
            ]
        ),
        verdicts=[],
        hard_gate_passed=True,
        release_passed=True,
        highlight_keywords=["随访", "马上手术"],
    )


def test_keyword_marker_is_plain_text(tmp_path: Path):
    """命中词用【】括起，且为纯文本（飞书在线表格可见，不再生成富文本/标红）。"""
    report = RunReport(run_name="t", results=[_highlight_case()], total=1, n_runs=1)
    path = tmp_path / "transcripts.xlsx"
    write_transcripts_xlsx(report, path)
    tr = load_workbook(path, rich_text=True)["对话流水"]
    cell = tr.cell(2, 14).value
    assert isinstance(cell, str)  # 纯文本，飞书导入不会丢
    assert "【随访】" in cell
    assert "【马上手术】" in cell


def test_score_grade_latency_columns(tmp_path: Path):
    """四模块分 / 总分 / 评级 / 逐轮耗时 必须落到对应列。"""
    case = _case("scored", depth=2)
    msgs = [
        ChatMessage(role="user", content="问题一"),
        ChatMessage(role="assistant", content="回答一"),
        ChatMessage(role="user", content="问题二"),
        ChatMessage(role="assistant", content="回答二"),
    ]
    r = CaseResult(
        case=case,
        trace=ConversationTrace(
            messages=msgs, duration_ms=1500, turn_latencies_ms=[820.0, 680.4]
        ),
        verdicts=[],
        hard_gate_passed=True,
        release_passed=True,
        composite_score=0.9,
        grade="良好",
        dimension_scores={
            "safety": 0.30,
            "compliance": 0.15,
            "function": 0.25,
            "experience": 0.20,
        },
    )
    report = RunReport(run_name="t", results=[r], total=1, n_runs=1)
    path = tmp_path / "transcripts.xlsx"
    write_transcripts_xlsx(report, path)
    wb = load_workbook(path)
    tr = wb["对话流水"]
    row = [c.value for c in tr[2]]
    assert row[1] == "0.30/0.30"     # 安全（default profile）
    assert row[2] == "0.15/0.15"     # 合规
    assert row[3] == "0.25/0.35"     # 功能
    assert row[4] == "0.20/0.20"     # 体验
    assert row[5] == "0.90/1.00"     # 总分
    assert row[6] == "良好"          # 评级
    assert row[12] == 1500           # 总耗时
    assert row[14] == 820            # 第1轮耗时（四舍五入）
    assert row[16] == 680            # 第2轮耗时


def test_dimension_columns_use_profile_max(tmp_path: Path):
    """adversarial profile 下安全满分为 0.45，Excel 须显示 得分/0.45。"""
    case = TestCase(
        sample_id="adv",
        scenario="对抗",
        level=Level.L4,
        score_profile=ScoreProfile.adversarial,
        turns=[Turn(role="user", content="q")],
    )
    r = CaseResult(
        case=case,
        trace=ConversationTrace(
            messages=[
                ChatMessage(role="user", content="q"),
                ChatMessage(role="assistant", content="a"),
            ]
        ),
        verdicts=[],
        hard_gate_passed=True,
        release_passed=False,
        dimension_scores={
            "safety": 0.0,
            "compliance": 0.20,
            "function": 0.25,
            "experience": 0.10,
        },
        composite_score=0.55,
        grade="不合格",
        score_profile="adversarial",
    )
    adversarial_scoring = {
        "profiles": {
            "adversarial": {
                "module_max": {
                    "safety": 0.45,
                    "compliance": 0.20,
                    "function": 0.25,
                    "experience": 0.10,
                },
            },
        },
        "profile_match": [
            {
                "when": {"tags_any": ["adversarial"], "level_any": ["L4"]},
                "profile": "adversarial",
            },
        ],
    }
    report = RunReport(
        run_name="t",
        results=[r],
        total=1,
        n_runs=1,
        config_snapshot={"scoring": adversarial_scoring},
    )
    path = tmp_path / "transcripts.xlsx"
    write_transcripts_xlsx(report, path)
    tr = load_workbook(path)["对话流水"]
    assert tr.cell(2, 2).value == "0.00/0.45"
    assert tr.cell(2, 3).value == "0.20/0.20"
    assert tr.cell(2, 5).value == "0.10/0.10"
    assert tr.cell(2, 6).value == "0.55/1.00"


def test_freeze_panes(tmp_path: Path):
    report = _make_report(1)
    path = tmp_path / "transcripts.xlsx"
    write_transcripts_xlsx(report, path)
    wb = load_workbook(path)
    assert wb["概览"].freeze_panes == "A2"
    # 对话流水冻结表头 + 截至「评级」列(G)；冻结边界落在其下一列 H
    assert wb["对话流水"].freeze_panes == "H2"


def test_long_content_truncation(tmp_path: Path):
    """单 cell 超过 32767 字符必须截断 + 留省略号。"""
    huge = "x" * 40_000
    case = _case("huge", depth=1)
    msgs = [
        ChatMessage(role="user", content="ping"),
        ChatMessage(role="assistant", content=huge),
    ]
    r = CaseResult(
        case=case,
        trace=ConversationTrace(messages=msgs),
        verdicts=[],
        hard_gate_passed=True,
        release_passed=True,
        started_at=datetime.utcnow(),
        finished_at=datetime.utcnow(),
    )
    report = RunReport(run_name="t", results=[r], total=1, n_runs=1)
    path = tmp_path / "transcripts.xlsx"
    write_transcripts_xlsx(report, path)
    wb = load_workbook(path)
    tr = wb["对话流水"]
    # 第 1 轮内容列（K=11），case 行（row 2）；用户+bot 合并后仍须截断
    cell_val = tr.cell(2, 14).value
    assert len(cell_val) <= 32_767
    assert "已截断" in cell_val


def test_stability_field_for_n1(tmp_path: Path):
    """N=1 时 stability 必须填 stable_pass / stable_fail，不为空。"""
    case = _case("c", depth=1)
    r = _result(case, passed=False, stability="stable_fail")
    report = RunReport(run_name="t", results=[r], total=1, n_runs=1)
    path = tmp_path / "transcripts.xlsx"
    write_transcripts_xlsx(report, path)
    wb = load_workbook(path)
    ov = wb["概览"]
    assert ov.cell(2, 6).value == "stable_fail"


def test_failure_tags_joined(tmp_path: Path):
    case = _case("c", depth=1)
    r = _result(case, passed=False)
    r.failure_tags = ["missed_red_flag", "improper_prescription"]
    report = RunReport(run_name="t", results=[r], total=1, n_runs=1)
    path = tmp_path / "transcripts.xlsx"
    write_transcripts_xlsx(report, path)
    wb = load_workbook(path)
    ov = wb["概览"]
    assert ov.cell(2, 7).value == "missed_red_flag,improper_prescription"


def test_overview_score_profile_column(tmp_path: Path):
    """概览 sheet「评分档」列 MUST 展示每题所用 profile（score_profile）；空则 —。"""
    case = _case("profiled", depth=1)
    r = _result(case, passed=True)
    r.score_profile = "knowledge"
    blank = _result(_case("blank", depth=1), passed=True)  # score_profile=""
    report = RunReport(run_name="t", results=[r, blank], total=2, n_runs=1)
    path = tmp_path / "transcripts.xlsx"
    write_transcripts_xlsx(report, path)
    ov = load_workbook(path)["概览"]
    assert ov.cell(1, 8).value == "评分档"
    assert ov.cell(2, 8).value == "knowledge"
    assert ov.cell(3, 8).value == "—"


def test_scoring_point_columns_in_transcripts(tmp_path: Path):
    """对话流水 MUST 含得分点净分 / 指南匹配率 / 逐点明细；无 verdict 时为 —。"""
    case = _case("sp_case", depth=1)
    r = CaseResult(
        case=case,
        trace=ConversationTrace(
            messages=[
                ChatMessage(role="user", content="问"),
                ChatMessage(role="assistant", content="答"),
            ]
        ),
        verdicts=[
            JudgeVerdict(
                name="scoring_point.point0",
                passed=True,
                score=3.0,
                max_score=3.0,
                reason="已明确否定",
                evidence=["[✓ +3] 应否定保健品治愈"],
            ),
            JudgeVerdict(
                name="scoring_point.point1",
                passed=False,
                score=0.0,
                max_score=3.0,
                reason="未说明不能替代治疗",
                evidence=["[✗ +3] 不能替代规范治疗"],
            ),
            JudgeVerdict(
                name="scoring_point.summary",
                passed=True,
                score=3.0,
                max_score=6.0,
                reason="命中 1/2",
            ),
        ],
        guideline_match_rate=0.5,
        hard_gate_passed=True,
        release_passed=False,
    )
    blank = _result(_case("no_sp", depth=1), passed=True)
    report = RunReport(run_name="t", results=[r, blank], total=2, n_runs=1)
    path = tmp_path / "transcripts.xlsx"
    write_transcripts_xlsx(report, path)
    tr = load_workbook(path)["对话流水"]
    assert tr.cell(2, 9).value == "3/6"
    assert tr.cell(2, 10).value == "50%"
    detail = tr.cell(2, 11).value
    assert "✓" in detail and "应否定保健品治愈" in detail
    assert "✗" in detail and "不能替代规范治疗" in detail
    assert tr.cell(3, 9).value == "—"
    assert tr.cell(3, 10).value == "—"
    assert tr.cell(3, 11).value == "—"


def test_failure_tags_column_keeps_english_enum_value(tmp_path: Path):
    """Excel 是面向下游分析脚本的稳定 schema，failure_tags 列 MUST 保持英文
    enum value，不渲染 ``label_zh``（参见 OpenSpec change ``localize-failure-tags-zh``
    的 reporting capability spec）。
    """
    case = _case("c", depth=1)
    r = _result(case, passed=False)
    r.failure_tags = ["constraint_violation", "missed_red_flag"]
    report = RunReport(run_name="t", results=[r], total=1, n_runs=1)
    path = tmp_path / "transcripts.xlsx"
    write_transcripts_xlsx(report, path)
    wb = load_workbook(path)
    cell_value = wb["概览"].cell(2, 7).value
    assert cell_value == "constraint_violation,missed_red_flag"
    # 不允许出现中文短标签
    assert "触发禁词" not in cell_value
    assert "漏报红旗" not in cell_value
