"""Transcripts.xlsx —— 完整对话流水的 Excel 导出（排版 / 写入层）。

参见 OpenSpec change ``add-transcript-excel-output``：评测报告里只贴 Top N 失败 case 的对话，
医学审阅者要追溯所有 case 的全文必须打开 transcripts.xlsx。

纯内容派生（截断 / 折行估算 / 八维、三端与指南得分）已拆到
``transcript_cells.py``（change ``2026-06-02-split-transcript-cells``）；本模块只负责
openpyxl 的 sheet / 列宽 / 行高 / 冻结窗格 / 样式写入。

文件结构（两个 sheet）：
  * Sheet 1 ``概览``：每行 1 个 case，列 sample_id / level / depth / scenario / passed / stability / failure_tags
  * Sheet 2 ``对话流水``：**每行 1 个 case** 的宽表。前缀列为
    测试内容、八维原始分与最终分、三端分、总分（满分40）、评级、指南逐项分、
    扣分原因、轮数与总耗时(ms)，
    其后按轮次展开 ``第N轮（用户+Bot）`` 与 ``第N轮耗时(ms)`` 成对的列。

依赖：openpyxl >= 3.1
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from ..models import CaseResult, RunReport
from ..evaluation import DIMENSION_LABELS, EvaluationDimension
from .transcript_cells import (
    _case_title,
    _deduction_text,
    _display_lines,
    _test_content_cell,
    _turn_cell,
    _turns,
    _user_turn_count,
)

# 对话流水宽表的固定前缀列（其后动态追加每轮 内容+耗时 列对）。
# 首列用 sub_scenario（测试内容）更直观；八维分 + 三端分 + 总分 + 评级 + 扣分原因
# 一行看全表现（飞书文档已不再贴失败明细，全在此 Excel 看）。
_PREFIX_HEADERS = [
    "测试内容",
    *[DIMENSION_LABELS[d] for d in EvaluationDimension],
    "医生端",
    "护士端",
    "患者端",
    "总分",
    "评级",
    "扣分原因",
    "指南评分",
    "轮数",
    "总耗时(ms)",
]
_REASON_HEADER = "扣分原因"
_SCORING_DETAIL_HEADER = "指南评分"

# 内容列宽（字符单位）与行高估算参数。把行高按文字量写进文件，
# 打开即可见全文，无需手动拉高（Excel 行高上限 409 pt）。
_CONTENT_COL_WIDTH = 70
_REASON_COL_WIDTH = 46
_SCORING_DETAIL_COL_WIDTH = 52
_LINE_PT = 15.0
_MAX_ROW_PT = 409.0


def _write_overview(ws, results: list[CaseResult]) -> None:
    headers = [
        "sample_id",
        "level",
        "depth",
        "scenario",
        "passed",
        "stability",
        "failure_tags",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for r in results:
        ws.append(
            [
                r.case.sample_id,
                r.case.level.value,
                _user_turn_count(r),
                r.case.scenario,
                r.release_passed,
                r.stability,
                ",".join(r.failure_tags),
            ]
        )

    widths = {"A": 32, "B": 8, "C": 8, "D": 18, "E": 10, "F": 14, "G": 40}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"


def _write_transcripts(ws, results: list[CaseResult]) -> None:
    """每行 1 个 case 的宽表：前缀摘要列 + 逐轮 (内容, 耗时) 列对。"""
    per_case_turns = [_turns(r) for r in results]
    max_turns = max((len(t) for t in per_case_turns), default=0)

    headers = list(_PREFIX_HEADERS)
    for t in range(1, max_turns + 1):
        headers.append(f"第{t}轮（用户+Bot）")
        headers.append(f"第{t}轮耗时(ms)")
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    wrap = Alignment(wrap_text=True, vertical="top")

    n_prefix = len(_PREFIX_HEADERS)
    reason_col = _PREFIX_HEADERS.index(_REASON_HEADER) + 1  # 1-based
    scoring_detail_col = _PREFIX_HEADERS.index(_SCORING_DETAIL_HEADER) + 1
    for r, turns in zip(results, per_case_turns):
        d = r.dimension_scores or {}
        raw = r.dimension_raw_scores or {}
        guideline_detail = "\n".join(
            f"{item.get('id')}: {item.get('score', 0):g}/{item.get('max_score', 0):g}"
            f" · {item.get('reason') or '—'}"
            for item in r.guideline_scores
        ) or "—"
        row: list = [
            _test_content_cell(r),
            *[
                (
                    f"{raw.get(dim.value, 0):g}→{d.get(dim.value, 0):g}/5"
                    if raw.get(dim.value) != d.get(dim.value)
                    else f"{d.get(dim.value, 0):g}/5"
                )
                for dim in EvaluationDimension
            ],
            f"{r.end_scores.get('doctor', 0):g}/15",
            f"{r.end_scores.get('nurse', 0):g}/15",
            f"{r.end_scores.get('user', 0):g}/15",
            f"{(r.composite_score or 0):g}/40",
            r.grade or "—",
            _deduction_text(r),
            guideline_detail,
            len(turns),
            r.trace.duration_ms,
        ]
        plain_texts: list[str] = []
        for user, bot, lat in turns:
            cell_value = _turn_cell(user, bot, [])
            plain_texts.append(cell_value)
            row.append(cell_value)
            row.append(round(lat) if lat is not None else "")
        ws.append(row)
        ws.cell(row=ws.max_row, column=1).alignment = wrap
        # 扣分原因 / 指南评分列换行
        ws.cell(row=ws.max_row, column=reason_col).alignment = wrap
        ws.cell(row=ws.max_row, column=scoring_detail_col).alignment = wrap
        # 内容列（前缀之后、每两列的第 1 列）换行
        for t in range(len(turns)):
            content_col = n_prefix + 1 + t * 2
            ws.cell(row=ws.max_row, column=content_col).alignment = wrap
        # 按该行最长内容（含测试内容/扣分原因列）估算并写入行高
        line_counts = [_display_lines(txt, _CONTENT_COL_WIDTH) for txt in plain_texts]
        title_cell = row[0]
        if isinstance(title_cell, str):
            line_counts.append(_display_lines(title_cell, 30))
        reason_text = row[reason_col - 1]
        if isinstance(reason_text, str):
            line_counts.append(_display_lines(reason_text, _REASON_COL_WIDTH))
        sp_detail_text = row[scoring_detail_col - 1]
        if isinstance(sp_detail_text, str) and sp_detail_text != "—":
            line_counts.append(
                _display_lines(sp_detail_text, _SCORING_DETAIL_COL_WIDTH)
            )
        if line_counts:
            ws.row_dimensions[ws.max_row].height = min(
                _MAX_ROW_PT, max(line_counts) * _LINE_PT + 4
            )

    # 列宽：测试内容/扣分原因/指南评分给宽些，其余分数与状态列紧凑。
    prefix_widths = [
        30,
        *([12] * 8),
        10,
        10,
        10,
        10,
        8,
        _REASON_COL_WIDTH,
        _SCORING_DETAIL_COL_WIDTH,
        6,
        12,
    ]
    for i, w in enumerate(prefix_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for t in range(max_turns):
        content_col = n_prefix + 1 + t * 2
        ws.column_dimensions[get_column_letter(content_col)].width = _CONTENT_COL_WIDTH
        ws.column_dimensions[get_column_letter(content_col + 1)].width = 12

    # 冻结表头行 + 截至「评级」列的身份/分数列；扣分原因、轮数、耗时及对话明细
    # 一起参与横向滚动，腾出屏宽看长对话明细，同时关键分级始终可见。
    freeze_col = _PREFIX_HEADERS.index("评级") + 2  # 评级列(1-based)的下一列
    ws.freeze_panes = f"{get_column_letter(freeze_col)}2"


def write_transcripts_xlsx(report: RunReport, path: Path) -> Path:
    """把 RunReport 的所有 case + trace 写为双 sheet xlsx。

    命中关键词统一用 ``【关键词】`` 纯文本标记（飞书在线表格可见）。
    """
    wb = Workbook()
    overview = wb.active
    overview.title = "概览"
    transcripts = wb.create_sheet("对话流水")

    _write_overview(overview, report.results)
    _write_transcripts(transcripts, report.results)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path
