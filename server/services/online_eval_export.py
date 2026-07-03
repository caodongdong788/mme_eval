"""线上评测清单导出：按筛选条件生成 xlsx 并发布为飞书表格。"""

from __future__ import annotations

import asyncio
import logging
import re
from dataclasses import dataclass
from datetime import datetime
from functools import cache
from io import BytesIO
from pathlib import Path
from typing import Any, Callable, Iterable, Optional

from fastapi import HTTPException
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.units import pixels_to_points
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from PIL import Image as PILImage
from sqlalchemy.orm import Session

from ..auth import SessionExpired, ensure_fresh_token
from ..feishu_drive import FeishuDriveError
from ..feishu_media import fetch_media
from ..models_db import FeishuUser, OnlineEval, OnlineEvalCase
from ..paths import safe_join
from ..settings import Settings, get_settings

log = logging.getLogger(__name__)

# 与 benchmarks._cell_text 生成的图片占位文本对齐：[图片：image_token=XXX，尺寸=WxH]
_IMAGE_RE = re.compile(r"\[图片[：:]\s*image_token=([A-Za-z0-9_-]+)(?:[，,]\s*尺寸=(\d+)x(\d+))?\]")
_IMAGE_DISPLAY_MAX_WIDTH = 360  # 展示宽度上限；原图数据不降采样，避免飞书内放大后发糊
_IMAGE_DISPLAY_MAX_HEIGHT = 560
_IMAGE_EXPORT_SCALE = 3  # 多图拼接时的导出倍率，显示 360px 时实际保留约 1080px
_IMAGE_GAP = 12  # 同格多图竖直堆叠间距（展示像素）
_TEXT_LINE_PX = 22
_TEXT_ROW_PADDING_PX = 14
_MAX_TEXT_ROW_HEIGHT_PT = 300
_MIN_DIALOGUE_COL_WIDTH = 32
_MAX_DIALOGUE_COL_WIDTH = 62

# token → 图片 bytes（None 表示下载失败）；导出时可传 None 表示不嵌图（未登录场景）。
ImageFetcher = Callable[[str], Optional[bytes]]
from .feishu_transcript_export import import_xlsx_as_sheet, publish_xlsx_to_lark
from .online_evals import (
    ROLE_LABELS,
    ROLE_ORDER,
    classify_missing_roles,
    get_online_eval_detail,
    _resolve_online_judge,
)

_ORDINAL_PREFIX = [
    "第一",
    "第二",
    "第三",
    "第四",
    "第五",
    "第六",
    "第七",
    "第八",
    "第九",
    "第十",
    "第十一",
    "第十二",
    "第十三",
    "第十四",
    "第十五",
    "第十六",
    "第十七",
    "第十八",
    "第十九",
    "第二十",
]


@dataclass(frozen=True)
class _CellImageRef:
    token: str
    placeholder: str


@dataclass(frozen=True)
class _PreparedImage:
    data: bytes
    display_width: int
    display_height: int


@dataclass(frozen=True)
class _CellPayload:
    text: str
    refs: list[_CellImageRef]
    fallback_text: str


def split_filter_values(raw: Optional[str]) -> list[str]:
    """解析前端传来的逗号分隔筛选值。"""
    if not raw:
        return []
    return [part.strip() for part in raw.split(",") if part.strip()]


def _score_matches(score: float, bucket: str) -> bool:
    if bucket == "gte40_5":
        return score >= 40.5
    if bucket == "36to40_5":
        return 36 <= score < 40.5
    if bucket == "27to36":
        return 27 <= score < 36
    if bucket == "lt27":
        return score < 27
    return False


def _matches_any_score_bucket(score: float, buckets: Iterable[str]) -> bool:
    selected = list(buckets)
    return not selected or any(_score_matches(score, bucket) for bucket in selected)


def filter_online_eval_cases(
    cases: list[OnlineEvalCase],
    *,
    gate_statuses: list[str] | None = None,
    score_buckets: list[str] | None = None,
    grades: list[str] | None = None,
) -> list[OnlineEvalCase]:
    """按详情表当前筛选条件过滤 case；列间 AND，列内多选 OR。"""
    gates = gate_statuses or []
    scores = score_buckets or []
    selected_grades = grades or []
    return [
        case
        for case in cases
        if (not gates or case.gate_status in gates)
        and _matches_any_score_bucket(case.total_score, scores)
        and (not selected_grades or case.grade in selected_grades)
    ]


def _round_prefix(index: int) -> str:
    if 1 <= index <= len(_ORDINAL_PREFIX):
        return _ORDINAL_PREFIX[index - 1]
    return f"第{index}"


def _append_assistant(turns: list[list[str]], content: str) -> None:
    if not turns:
        turns.append(["", content])
        return
    if turns[-1][1]:
        turns[-1][1] = f"{turns[-1][1]}\n{content}"
    else:
        turns[-1][1] = content


def case_dialogue_turns(case: OnlineEvalCase) -> list[tuple[str, str]]:
    """将 raw_messages 还原成按轮次排列的「用户输入 / Cx 输出」。"""
    turns: list[list[str]] = []
    raw_messages = case.raw_messages if isinstance(case.raw_messages, list) else []
    for msg in raw_messages:
        if not isinstance(msg, dict):
            continue
        role = str(msg.get("role") or "").strip()
        content = str(msg.get("content") or "").strip()
        if not content:
            continue
        if role == "user":
            turns.append([content, ""])
        elif role in {"assistant", "bot", "cx"}:
            _append_assistant(turns, content)

    if not turns:
        user_text = (case.user_text or "").strip()
        assistant_text = (case.assistant_text or "").strip()
        if user_text or assistant_text:
            turns.append([user_text, assistant_text])

    return [(user, assistant) for user, assistant in turns]


def _headers(max_turns: int, split_user_by_turn: list[bool] | None = None) -> list[str]:
    split_user_by_turn = split_user_by_turn or [False] * max_turns
    headers: list[str] = []
    for index in range(1, max_turns + 1):
        prefix = _round_prefix(index)
        headers.append(f"{prefix}轮用户输入")
        if index - 1 < len(split_user_by_turn) and split_user_by_turn[index - 1]:
            headers.append(f"{prefix}轮用户输入")
        headers.append(f"{prefix}轮Cx输出")
    return headers


def _extract_cell_images(text: str) -> tuple[str, list[_CellImageRef]]:
    """从单元格文本拆出图片引用，并返回去掉图片占位后的纯文本。"""
    refs = [
        _CellImageRef(
            token=match.group(1),
            placeholder=match.group(0),
        )
        for match in _IMAGE_RE.finditer(text)
    ]
    if not refs:
        return text, []
    clean = _IMAGE_RE.sub("", text)
    # 去掉因删除占位产生的多余空行/首尾空白。
    clean = "\n".join(line.strip() for line in clean.splitlines() if line.strip())
    return clean, refs


def _image_placeholder_text(refs: list[_CellImageRef]) -> str:
    return "\n".join(ref.placeholder for ref in refs)


def _cell_payload(raw: Any) -> _CellPayload:
    text = str(raw or "")
    clean, refs = _extract_cell_images(text)
    if not refs:
        return _CellPayload(text=text, refs=[], fallback_text=text)
    return _CellPayload(text=clean, refs=refs, fallback_text=text)


def _split_user_input_payload(raw: Any) -> tuple[_CellPayload, _CellPayload]:
    payload = _cell_payload(raw)
    if not payload.refs:
        return payload, _CellPayload(text="", refs=[], fallback_text="")
    return (
        _CellPayload(text=payload.text, refs=[], fallback_text=payload.text),
        _CellPayload(
            text="",
            refs=payload.refs,
            fallback_text=_image_placeholder_text(payload.refs),
        ),
    )


def _needs_split_user_input(text: str) -> bool:
    clean, refs = _extract_cell_images(text)
    return bool(clean.strip() and refs)


def _split_user_columns_by_turn(
    rows: list[tuple[OnlineEvalCase, list[tuple[str, str]]]],
    max_turns: int,
) -> list[bool]:
    split = [False] * max_turns
    for _case, turns in rows:
        for index, (user_text, _assistant_text) in enumerate(turns[:max_turns]):
            if _needs_split_user_input(user_text):
                split[index] = True
    return split


def _fit_image_display(width: int, height: int) -> tuple[int, int]:
    """按展示上限等比计算图片显示尺寸，不放大小图。"""
    if width <= 0 or height <= 0:
        return _IMAGE_DISPLAY_MAX_WIDTH, _IMAGE_DISPLAY_MAX_WIDTH
    scale = min(
        1.0,
        _IMAGE_DISPLAY_MAX_WIDTH / width,
        _IMAGE_DISPLAY_MAX_HEIGHT / height,
    )
    display_width = max(1, round(width * scale))
    display_height = max(1, round(height * scale))
    return display_width, display_height


def _prepare_single_image(blob: bytes) -> _PreparedImage:
    """单图直接保留原始 bytes，只调整 Excel 展示尺寸。"""
    img = PILImage.open(BytesIO(blob))
    img.load()
    display_width, display_height = _fit_image_display(img.width, img.height)
    return _PreparedImage(blob, display_width, display_height)


def _stack_images(blobs: list[bytes]) -> _PreparedImage:
    """多图拼接成一张高分辨率 PNG；返回数据与展示尺寸。"""
    frames: list[PILImage.Image] = []
    for blob in blobs:
        img = PILImage.open(BytesIO(blob))
        img.load()
        if img.mode not in ("RGB", "RGBA"):
            img = img.convert("RGB")
        frames.append(img)

    max_width = max((frame.width for frame in frames), default=_IMAGE_DISPLAY_MAX_WIDTH)
    max_height = max((frame.height for frame in frames), default=_IMAGE_DISPLAY_MAX_WIDTH)
    display_width, _display_height = _fit_image_display(max_width, max_height)
    export_width = min(max_width, max(1, display_width * _IMAGE_EXPORT_SCALE))
    scale = export_width / display_width if display_width else 1
    export_gap = max(1, round(_IMAGE_GAP * scale))

    resized: list[PILImage.Image] = []
    for frame in frames:
        if frame.width != export_width:
            new_h = max(1, round(frame.height * export_width / frame.width))
            frame = frame.resize((export_width, new_h), PILImage.Resampling.LANCZOS)
        resized.append(frame)

    total_h = sum(f.height for f in resized) + export_gap * (len(resized) - 1)
    canvas = PILImage.new("RGB", (export_width, total_h), "white")
    offset = 0
    for frame in resized:
        canvas.paste(frame, (0, offset), frame if frame.mode == "RGBA" else None)
        offset += frame.height + export_gap

    buffer = BytesIO()
    canvas.save(buffer, format="PNG")
    return _PreparedImage(
        data=buffer.getvalue(),
        display_width=display_width,
        display_height=max(1, round(total_h / scale)),
    )


def _prepare_cell_image(blobs: list[bytes]) -> _PreparedImage:
    if len(blobs) == 1:
        return _prepare_single_image(blobs[0])
    return _stack_images(blobs)


def _text_units(text: str) -> int:
    """估算 Excel 列宽单位；中文等宽按 2 个英文字符粗略计算。"""
    return sum(2 if ord(char) > 127 else 1 for char in text)


def _max_line_units(text: str) -> int:
    return max((_text_units(line) for line in str(text or "").splitlines()), default=0)


def _estimate_text_lines(text: str, column_width: float) -> int:
    if not text:
        return 1
    usable_width = max(8, int(column_width) - 2)
    total = 0
    for line in str(text).splitlines() or [""]:
        total += max(1, (_text_units(line) + usable_width - 1) // usable_width)
    return total


def _display_width_to_column_width(display_width: int) -> float:
    # Excel 列宽近似：1 个宽度单位约 7px；额外留一点水平边距。
    return min(_MAX_DIALOGUE_COL_WIDTH, max(_MIN_DIALOGUE_COL_WIDTH, (display_width + 18) / 7))


def _row_height_for_text(texts: Iterable[str], widths: list[float]) -> float:
    max_lines = 1
    for index, text in enumerate(texts):
        width = widths[index] if index < len(widths) else _MIN_DIALOGUE_COL_WIDTH
        max_lines = max(max_lines, _estimate_text_lines(text, width))
    return min(
        _MAX_TEXT_ROW_HEIGHT_PT,
        pixels_to_points(max_lines * _TEXT_LINE_PX + _TEXT_ROW_PADDING_PX),
    )


def _build_image_fetcher(access_token: str) -> ImageFetcher:
    """构造带缓存的图片下载器；下载失败返回 None（由调用方回退文本）。"""

    @cache
    def fetch(token: str) -> Optional[bytes]:
        try:
            return fetch_media(access_token, token).content
        except Exception as exc:  # noqa: BLE001 - 图片失败不应阻断导出
            log.warning("导出清单下载飞书图片失败 token=%s：%s", token, exc)
            return None

    return fetch


def _fill_sheet(
    ws: Worksheet,
    rows: list[tuple[OnlineEvalCase, list[tuple[str, str]]]],
    max_turns: int,
    split_user_by_turn: list[bool],
    image_fetcher: Optional[ImageFetcher],
) -> None:
    """把一组 case 写入单个 sheet；含图单元格下载并嵌入真实图片，套用统一表头样式。"""
    headers = _headers(max_turns, split_user_by_turn)
    ws.append(headers)
    column_widths = [_MIN_DIALOGUE_COL_WIDTH] * len(headers)

    for case, turns in rows:
        payloads: list[_CellPayload] = []
        for index in range(max_turns):
            if index < len(turns):
                user_text, assistant_text = turns[index]
                if split_user_by_turn[index]:
                    user_text_payload, user_image_payload = _split_user_input_payload(user_text)
                    payloads.extend([
                        user_text_payload,
                        user_image_payload,
                        _cell_payload(assistant_text),
                    ])
                else:
                    payloads.extend([_cell_payload(user_text), _cell_payload(assistant_text)])
            else:
                if split_user_by_turn[index]:
                    payloads.extend([
                        _CellPayload(text="", refs=[], fallback_text=""),
                        _CellPayload(text="", refs=[], fallback_text=""),
                        _CellPayload(text="", refs=[], fallback_text=""),
                    ])
                else:
                    payloads.extend([
                        _CellPayload(text="", refs=[], fallback_text=""),
                        _CellPayload(text="", refs=[], fallback_text=""),
                    ])

        values: list[Any] = [
            payload.text if image_fetcher is not None else payload.fallback_text
            for payload in payloads
        ]

        # 先解析图片、替换为纯文本，收集待嵌图片；无 fetcher 时保留原文本。
        pending: list[tuple[int, _PreparedImage]] = []  # (列 0-based, image)
        if image_fetcher is not None:
            for col, payload in enumerate(payloads):
                if not payload.refs:
                    continue
                blobs = [data for ref in payload.refs if (data := image_fetcher(ref.token))]
                if not blobs:
                    values[col] = payload.fallback_text
                    continue  # 全部下载失败 → 保留原占位文本
                prepared = _prepare_cell_image(blobs)
                values[col] = payload.text
                pending.append((col, prepared))
                column_widths[col] = max(
                    column_widths[col],
                    _display_width_to_column_width(prepared.display_width),
                )

        for col, raw in enumerate(values):
            column_widths[col] = min(
                _MAX_DIALOGUE_COL_WIDTH,
                max(column_widths[col], min(_MAX_DIALOGUE_COL_WIDTH, _max_line_units(str(raw)) + 4)),
            )

        ws.append(values)
        row_idx = ws.max_row
        for col, prepared in pending:
            image = XLImage(BytesIO(prepared.data))
            image.width = prepared.display_width
            image.height = prepared.display_height
            ws.add_image(image, f"{get_column_letter(col + 1)}{row_idx}")
            if values[col]:
                # 图片浮在单元格上方；给文本预留垂直空白，避免导入飞书后重叠。
                blank_lines = max(1, round(prepared.display_height / _TEXT_LINE_PX))
                values[col] = "\n" * blank_lines + str(values[col])
                ws.cell(row=row_idx, column=col + 1).value = values[col]
        row_height = _row_height_for_text([str(value or "") for value in values], column_widths)
        if pending:
            image_height = max(image.display_height for _col, image in pending)
            row_height = max(row_height, pixels_to_points(image_height + _TEXT_ROW_PADDING_PX))
        ws.row_dimensions[row_idx].height = row_height
        for cell in ws[row_idx]:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    header_fill = PatternFill(fill_type="solid", fgColor="EAF2FF")
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    for index, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width


def _write_cases_xlsx(
    cases: list[OnlineEvalCase],
    xlsx_path: Path,
    image_fetcher: Optional[ImageFetcher] = None,
) -> None:
    """按复核角色（医生/护士/患者）拆分为多个 sheet；列布局全局统一。"""
    rows = [(case, case_dialogue_turns(case)) for case in cases]
    max_turns = max((len(turns) for _case, turns in rows), default=1)

    # 按角色分组；未分类（理论上分类后不会出现）兜底归患者。
    grouped: dict[str, list[tuple[OnlineEvalCase, list[tuple[str, str]]]]] = {
        role: [] for role in ROLE_ORDER
    }
    for case, turns in rows:
        role = case.review_role if case.review_role in grouped else "patient"
        grouped[role].append((case, turns))

    wb = Workbook()
    wb.remove(wb.active)  # 丢弃默认空 sheet，只保留有 case 的角色 sheet
    for role in ROLE_ORDER:
        role_rows = grouped[role]
        if not role_rows:
            continue
        ws = wb.create_sheet(ROLE_LABELS[role])
        split_user_by_turn = _split_user_columns_by_turn(role_rows, max_turns)
        _fill_sheet(ws, role_rows, max_turns, split_user_by_turn, image_fetcher)
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)


def _ensure_review_roles(
    session: Session,
    row: OnlineEval,
    cases: list[OnlineEvalCase],
) -> None:
    """对未分类 case 用 LLM 判定复核角色并落库（缓存，避免重复导出重复调用）。"""
    missing = [case for case in cases if not case.review_role]
    if not missing:
        return
    judge = _resolve_online_judge(session, row.judge_model_id)
    roles = asyncio.run(classify_missing_roles(missing, judge))
    for case in missing:
        case.review_role = roles.get(case.id, "patient")
    # 立即提交：即使随后飞书上传失败，分类缓存也已落库，重试不再重复调用 LLM。
    session.commit()


def export_online_eval_cases(
    session: Session,
    eval_id: int,
    *,
    gate_statuses: list[str] | None = None,
    score_buckets: list[str] | None = None,
    grades: list[str] | None = None,
    parent_folder_token: Optional[str] = None,
    current_user: Optional[FeishuUser] = None,
    settings: Optional[Settings] = None,
) -> dict[str, Any]:
    settings = settings or get_settings()
    row = get_online_eval_detail(session, eval_id)
    cases = filter_online_eval_cases(
        row.cases,
        gate_statuses=gate_statuses,
        score_buckets=score_buckets,
        grades=grades,
    )
    if not cases:
        raise HTTPException(status_code=400, detail="当前过滤条件下没有可导出的线上评测 case")

    _ensure_review_roles(session, row, cases)

    # 登录用户先刷新 token，再据此构造图片下载器，导出时把图片占位还原为真实嵌图；
    # 未登录（lark-cli 分支）无 token，image_fetcher=None，图片保留原文本兜底。
    image_fetcher: Optional[ImageFetcher] = None
    if current_user is not None:
        try:
            ensure_fresh_token(session, current_user, settings)
        except SessionExpired:
            raise HTTPException(status_code=401, detail="飞书会话已过期，请重新登录")
        image_fetcher = _build_image_fetcher(current_user.access_token)

    try:
        out_dir = safe_join(settings.outputs_dir, "online_eval_exports")
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="非法的导出目录") from exc
    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    xlsx_path = out_dir / f"online_eval_{eval_id}_cases_{timestamp}.xlsx"
    _write_cases_xlsx(cases, xlsx_path, image_fetcher)

    token = "" if parent_folder_token is None else parent_folder_token
    title = f"{row.name or f'线上评测_{eval_id}'}_评测清单"

    if current_user is not None:
        try:
            url = import_xlsx_as_sheet(
                current_user.access_token,
                xlsx_path,
                folder_token=token,
                title=title,
            )
        except FeishuDriveError as exc:
            raise HTTPException(
                status_code=502,
                detail=(
                    f"飞书导出失败：{exc}。请确认：①已开通 drive:drive 权限；"
                    "②当前账号对目标文件夹有写权限；③可留空 token 改为个人根目录。"
                ),
            )
        return {"url": url, "count": len(cases), "filename": xlsx_path.name}

    url = publish_xlsx_to_lark(xlsx_path, parent_folder_token=token, title=title)
    if not url:
        raise HTTPException(
            status_code=502,
            detail=(
                "飞书发布失败。请确认已安装并登录 lark-cli，或先用飞书账号登录平台后重试。"
            ),
        )
    return {"url": url, "count": len(cases), "filename": xlsx_path.name}
